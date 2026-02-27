#!/usr/bin/env python3
"""Import Job_Search_Tracker.xlsx into DynamoDB.

Usage:
    uv run scripts/import_excel.py path/to/Job_Search_Tracker.xlsx [--table jobtracker-dev] [--dry-run]
"""

import argparse
import sys
from pathlib import Path

import boto3
import openpyxl
from ulid import ULID


def import_excel(filepath, table_name, dry_run=False):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Applications"]

    print(f"Reading from: {filepath}")
    print(f"Sheet: {ws.title}, Rows: {ws.max_row - 1}")
    print(f"Target table: {table_name}")
    if dry_run:
        print("DRY RUN — no data will be written")
    print()

    dynamodb = boto3.resource("dynamodb")
    table = dynamodb.Table(table_name)

    imported = 0
    skipped = 0
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        company = row[1].value
        if not company:
            skipped += 1
            continue

        company = str(company).strip()

        # Parse date
        date_val = row[0].value
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val:
            date_str = str(date_val)[:10]
        else:
            date_str = ""

        role = str(row[2].value or "").strip()
        level = str(row[3].value or "").strip()
        status = str(row[4].value or "Applied").strip()
        source = str(row[5].value or "").strip()
        link = str(row[6].value or "").strip()
        closed_val = row[7].value
        original_week = str(row[8].value or "").strip()
        notes = str(row[9].value or "").strip()

        # Fix data entry errors: URLs in status column
        if status.startswith("http"):
            status = "Applied"

        # Dedup check
        dedup_key = (date_str, company, role)
        if dedup_key in seen:
            skipped += 1
            continue
        seen.add(dedup_key)

        item = {
            "pk": "APPLICATION",
            "sk": str(ULID()),
            "date": date_str,
            "company": company,
            "role": role,
        }

        # Only add non-empty optional fields
        if level and level != "None":
            item["level"] = level
        if status and status != "None":
            item["status"] = status
        if source and source != "None":
            item["source"] = source
        if link and link != "None":
            item["link"] = link
        if closed_val:
            item["closed"] = True
        if original_week and original_week != "None":
            item["original_week"] = original_week
        if notes and notes != "None":
            item["notes"] = notes

        if dry_run:
            if imported < 5:
                print(f"  Would import: {date_str} | {company} | {role} | {status}")
        else:
            table.put_item(Item=item)

        imported += 1

    print()
    print(f"{'Would import' if dry_run else 'Imported'}: {imported}")
    print(f"Skipped: {skipped}")
    return imported, skipped


def main():
    parser = argparse.ArgumentParser(description="Import Excel tracker into DynamoDB")
    parser.add_argument("file", help="Path to Job_Search_Tracker.xlsx")
    parser.add_argument("--table", default="jobtracker-dev", help="DynamoDB table name")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"File not found: {args.file}")
        sys.exit(1)

    import_excel(args.file, args.table, args.dry_run)


if __name__ == "__main__":
    main()
